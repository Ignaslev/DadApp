"""
Importuoja duomenis iš Excel failo į Django sistemą.
Naudojimas: python manage.py import_excel "kelias/iki/failo.xlsm"
"""
import datetime
from datetime import date
from decimal import Decimal

import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction

from inventory.models import Client, Invoice, InvoiceLine, Payment, Product, Purchase, Sale


class Command(BaseCommand):
    help = 'Importuoja duomenis iš Excel failo'

    def add_arguments(self, parser):
        parser.add_argument('excel_path', type=str)
        parser.add_argument('--skip-clients', action='store_true')
        parser.add_argument('--skip-products', action='store_true')
        parser.add_argument('--skip-purchases', action='store_true')
        parser.add_argument('--skip-sales', action='store_true')
        parser.add_argument('--skip-payments', action='store_true')
        parser.add_argument(
            '--history-updates-stock',
            action='store_true',
            help='Pirkimų importas atnaujina sandėlio likučius iš istorijos. Nenaudokite kartu su SANDELIS snapshot, jei norite išvengti dvigubo skaičiavimo.',
        )

    def handle(self, *args, **options):
        path = options['excel_path']
        self.stdout.write(f'\nAtidaromas: {path}')
        self.stdout.write('Kraunama...\n')
        try:
            wb = openpyxl.load_workbook(path, keep_vba=True, data_only=True)
        except Exception as e:
            self.stderr.write(f'Klaida: {e}')
            return

        if not options['skip_products'] and not options['history_updates_stock']:
            self.stdout.write('Naudojamas SANDELIS snapshot režimas: istorija bus importuota ataskaitoms, bet nepakeis esamų likučių.')

        with transaction.atomic():
            if not options['skip_clients']:
                self._import_clients(wb)
            if not options['skip_products']:
                self._import_products(wb)
            if not options['skip_purchases']:
                self._import_purchases(wb, update_stock=options['history_updates_stock'])
            if not options['skip_sales']:
                self._import_sales(wb)
            if not options['skip_payments']:
                self._import_payments(wb)

        self.stdout.write(self.style.SUCCESS('\nImportas baigtas!\n'))

    def _dec(self, val, default=0):
        try:
            if val is None or val == '':
                return Decimal(str(default))
            return Decimal(str(round(float(val), 8)))
        except Exception:
            return Decimal(str(default))

    def _date(self, val):
        if isinstance(val, datetime.datetime):
            return val.date()
        if isinstance(val, date):
            return val
        return None

    def _s(self, val):
        return '' if val is None else str(val).strip()

    def _vat_percent(self, raw, default='21'):
        try:
            value = float(raw)
            value = value * 100 if value <= 1 else value
            return Decimal(str(round(value, 2)))
        except Exception:
            return Decimal(str(default))

    def _client_code_from_name(self, name):
        return name[:48].replace('"', '').replace(' ', '_')

    def _import_clients(self, wb):
        if 'KLIENTAI' not in wb.sheetnames:
            self.stdout.write('  [KLIENTAI] nerastas')
            return
        ws = wb['KLIENTAI']
        created = skipped = 0
        for row in ws.iter_rows(min_row=5, values_only=True):
            name = self._s(row[2]) if len(row) > 2 else ''
            if not name:
                continue
            code = self._s(row[5]) if len(row) > 5 else ''
            if not code:
                code = self._client_code_from_name(name)
            if not code:
                continue
            base_code = code[:48]
            code = base_code
            suffix = 1
            while Client.objects.filter(code=code).exclude(name=name).exists():
                code = f"{base_code[:45]}_{suffix}"
                suffix += 1
            _, was_created = Client.objects.get_or_create(
                code=code,
                defaults={
                    'name': name,
                    'country': self._s(row[3]) if len(row) > 3 else '',
                    'country_code': self._s(row[4]) if len(row) > 4 else '',
                    'vat_code': self._s(row[6]) if len(row) > 6 else '',
                    'address': self._s(row[7]) if len(row) > 7 else '',
                    'contacts': self._s(row[8]) if len(row) > 8 else '',
                    'notes': self._s(row[9]) if len(row) > 9 else '',
                },
            )
            if was_created:
                created += 1
            else:
                skipped += 1
        self.stdout.write(f'  [KLIENTAI]    sukurta: {created}, jau buvo: {skipped}')

    def _import_products(self, wb):
        if 'SANDELIS' not in wb.sheetnames:
            self.stdout.write('  [SANDELIS] nerastas')
            return
        ws = wb['SANDELIS']
        created = skipped = 0
        for row in ws.iter_rows(min_row=11, values_only=True):
            if not row or len(row) < 4:
                continue
            code = self._s(row[3])
            if not code or code == 'Kodas':
                continue
            desc1 = self._s(row[5]) if len(row) > 5 else ''
            if not desc1:
                desc1 = self._s(row[1]) if len(row) > 1 else code
            qty = self._dec(row[17] if len(row) > 17 else 0)
            product_defaults = {
                'product_type': self._s(row[4]) if len(row) > 4 else 'Prekės',
                'description1': desc1,
                'description2': self._s(row[6]) if len(row) > 6 else '',
                'description3': self._s(row[7]) if len(row) > 7 else '',
                'description4': self._s(row[8]) if len(row) > 8 else '',
                'unit': self._s(row[10]) if len(row) > 10 else 'vnt',
                'quantity': qty,
                'avg_cost': self._dec(row[11] if len(row) > 11 else 0),
                'sale_price': self._dec(row[18] if len(row) > 18 else 0),
                'weight_kg': self._dec(row[20] if len(row) > 20 else 0),
                'sold_quantity': self._dec(row[14] if len(row) > 14 else 0),
                'consignment_quantity': self._dec(row[15] if len(row) > 15 else 0),
                'stock_adjustment': self._dec(row[16] if len(row) > 16 else 0),
                'package_code': self._s(row[19]) if len(row) > 19 else '',
                'notes': self._s(row[23]) if len(row) > 23 else '',
            }
            _, was_created = Product.objects.get_or_create(code=code, defaults=product_defaults)
            if was_created:
                created += 1
            else:
                skipped += 1
        self.stdout.write(f'  [SANDELIS]    sukurta: {created}, jau buvo: {skipped}')

    def _import_purchases(self, wb, update_stock=False):
        if 'PIRKIMAI' not in wb.sheetnames:
            self.stdout.write('  [PIRKIMAI] nerastas')
            return
        ws = wb['PIRKIMAI']
        created = skipped = errors = 0
        for row in ws.iter_rows(min_row=11, values_only=True):
            if not row or len(row) < 15:
                continue
            purchase_type = self._s(row[2])
            inv_num = self._s(row[3])
            inv_date = self._date(row[4])
            if not inv_num or not inv_date:
                continue
            supplier = self._s(row[6]) or '—'
            desc = self._s(row[7]) or self._s(row[22]) or inv_num
            product_code = self._s(row[11])
            quantity = self._dec(row[13])
            unit_price = self._dec(row[14])
            vat_rate = self._vat_percent(row[16], default='0')
            product = None
            if product_code:
                product = Product.objects.filter(code=product_code).first()
            try:
                existing = Purchase.objects.filter(
                    invoice_number=inv_num,
                    invoice_date=inv_date,
                    supplier=supplier,
                    description=desc,
                    quantity=quantity,
                    unit_price=unit_price,
                ).first()
                if existing:
                    skipped += 1
                    continue
                purchase = Purchase(
                    invoice_number=inv_num,
                    invoice_date=inv_date,
                    due_date=self._date(row[5]) if len(row) > 5 else None,
                    supplier=supplier,
                    purchase_type=purchase_type or 'Prekės',
                    product=product,
                    description=desc,
                    unit=self._s(row[12]) if len(row) > 12 else 'vnt',
                    quantity=quantity,
                    unit_price=unit_price,
                    vat_rate=vat_rate,
                    notes=f'Importuota iš Excel. Kodas: {product_code}' if product_code else 'Importuota iš Excel',
                )
                purchase.save(skip_stock_update=not update_stock)
                created += 1
            except Exception as exc:
                errors += 1
                self.stderr.write(f'    PIRKIMAI klaida {inv_num}: {exc}')
        self.stdout.write(f'  [PIRKIMAI]    sukurta: {created}, jau buvo: {skipped}, klaidos: {errors}')

    def _import_sales(self, wb):
        if 'PARDAVIMAI' not in wb.sheetnames:
            self.stdout.write('  [PARDAVIMAI] nerastas')
            return
        ws = wb['PARDAVIMAI']

        invoice_rows = {}
        for row in ws.iter_rows(min_row=6, values_only=True):
            if not row or len(row) < 14:
                continue
            inv_num = self._s(row[3])
            if not inv_num:
                continue
            invoice_rows.setdefault(inv_num, []).append(row)

        created_inv = skipped_inv = created_lines = errors = 0
        for inv_num, rows in invoice_rows.items():
            if Invoice.objects.filter(number=inv_num).exists():
                skipped_inv += 1
                continue

            first = rows[0]
            inv_date = self._date(first[4])
            if not inv_date:
                continue

            client_name = ' '.join(self._s(first[5]).split())
            if not client_name:
                continue

            client = (
                Client.objects.filter(name__iexact=client_name).first()
                or Client.objects.filter(name__icontains=client_name[:25]).first()
            )
            if not client:
                code = self._client_code_from_name(client_name)
                if Client.objects.filter(code=code).exists():
                    code = code[:44] + f'_{Client.objects.count()}'
                client = Client.objects.create(name=client_name, code=code)

            try:
                vat_rate = Decimal('21')
                due_date = None
                payment_term = first[6] if len(first) > 6 else None
                if isinstance(payment_term, (int, float)):
                    due_date = inv_date + datetime.timedelta(days=int(payment_term))
                else:
                    due_date = self._date(payment_term)

                for r in rows:
                    vr = r[17] if len(r) > 17 else None
                    if vr is not None:
                        vat_rate = self._vat_percent(vr, default='21')
                        break

                inv_type = 'PVM' if vat_rate > 0 else 'BE_PVM'
                invoice = Invoice.objects.create(
                    number=inv_num,
                    date=inv_date,
                    due_date=due_date,
                    client=client,
                    invoice_type=inv_type,
                    status='ISSUED',
                )
                created_inv += 1

                actual_line_number = 0
                for row in rows:
                    desc = self._s(row[7])
                    product_code = self._s(row[8])
                    if not desc and not product_code:
                        continue

                    product = Product.objects.filter(code=product_code).first() if product_code else None
                    qty = self._dec(row[12])
                    unit_price = self._dec(row[13])
                    cost_price = self._dec(row[11])
                    unit = self._s(row[9]) or (product.unit if product else 'vnt')
                    weight = self._dec(row[22] if len(row) > 22 else 0)
                    line_vat = self._vat_percent(row[17] if len(row) > 17 else 0.21, default='21')

                    actual_line_number += 1
                    line = InvoiceLine.objects.create(
                        invoice=invoice,
                        line_number=actual_line_number,
                        product=product,
                        description=desc or product_code or '—',
                        unit=unit,
                        quantity=qty,
                        unit_price=unit_price,
                        vat_rate=line_vat,
                        weight_kg=weight,
                    )
                    created_lines += 1

                    if qty > 0:
                        Sale.objects.create(
                            invoice=invoice,
                            invoice_line=line,
                            product=product,
                            date=inv_date,
                            client=client,
                            quantity=qty,
                            unit_price=unit_price,
                            cost_price=cost_price,
                        )
                invoice.update_status()
            except Exception as exc:
                errors += 1
                self.stderr.write(f'    Klaida {inv_num}: {exc}')

        self.stdout.write(
            f'  [PARDAVIMAI]  sąskaitos: {created_inv}, praleista: {skipped_inv}, eilutės: {created_lines}, klaidos: {errors}'
        )

    def _import_payments(self, wb):
        if 'APMOKEJIMAI' not in wb.sheetnames:
            self.stdout.write('  [APMOKEJIMAI] nerastas')
            return
        ws = wb['APMOKEJIMAI']
        created = skipped = not_found = 0
        for row in ws.iter_rows(min_row=5, values_only=True):
            if not row or len(row) < 6:
                continue
            inv_num = self._s(row[2])
            if not inv_num:
                continue
            paid = self._dec(row[5])
            if paid <= 0:
                continue
            try:
                invoice = Invoice.objects.get(number=inv_num)
            except Invoice.DoesNotExist:
                not_found += 1
                continue
            if Payment.objects.filter(invoice=invoice, amount=paid).exists():
                skipped += 1
                continue
            pay_date = self._date(row[9]) if len(row) > 9 else None
            status_str = self._s(row[7])
            Payment.objects.create(
                invoice=invoice,
                date=pay_date or invoice.date,
                amount=paid,
                method='TRANSFER',
                notes=f'Importuota iš Excel. Būsena: {status_str}' if status_str else 'Importuota iš Excel',
            )
            invoice.update_status()
            created += 1
        self.stdout.write(f'  [APMOKEJIMAI] sukurta: {created}, jau buvo: {skipped}, nerasta: {not_found}')
